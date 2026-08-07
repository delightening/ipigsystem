"""print-pdf FastAPI app — HTML→PDF via Jinja2 + WeasyPrint.

取代 `services/word-convert` daemon + `services/gotenberg` + `pdf-service`
docx/Chromium 路徑（自帶 WeasyPrint，無 Office COM / 無 LibreOffice / 無 Chromium）。

schemas/ + adapters/ 從 `pdf-service/app/` 複製，shape 完全對齊，
讓 backend 不需改 payload 即可切換到本服務。
"""
from __future__ import annotations

import asyncio
import base64
import hmac
import io
import logging
import os
import time
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request, status
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from jinja2 import Environment, FileSystemLoader, select_autoescape
from pydantic import BaseModel

from adapters import (
    audit_log as audit_log_ad,
    aup_protocol as aup_ad,
    blood_test as blood_test_ad,
    medical_record as medical_record_ad,
    review_reply as review_reply_ad,
    review_result as review_result_ad,
    surgery as surgery_ad,
    vet_patrol as vet_patrol_ad,
    vet_patrol_report as vet_patrol_report_ad,
    warehouse as warehouse_ad,
)
from floor_plan import render_floor_plan_png
from samples import SAMPLES
from schemas import (
    AuditLogPayload,
    AupProtocolPayload,
    BloodTestPayload,
    MedicalRecordPayload,
    PigApprovalPayload,
    ReviewReplyPayload,
    ReviewResultPayload,
    SurgeryPayload,
    VetPatrolPayload,
    VetPatrolReportPayload,
    WarehouseReportPayload,
)

log = logging.getLogger("print-pdf")
logging.basicConfig(level=logging.INFO)

BASE = Path(__file__).parent
TEMPLATES_DIR = BASE / "templates"
STATIC_DIR = BASE / "static"
TZ = ZoneInfo(os.environ.get("TZ", "Asia/Taipei"))


def _load_internal_token() -> tuple[str, bool]:
    """Read PDF_SERVICE_TOKEN from env var or PDF_SERVICE_TOKEN_FILE secret file.

    Returns ``(token, source_configured)``. ``source_configured`` is True when a
    token *source* was provided (non-empty env var, or a file path) — i.e. the
    operator intended auth to be ON. Backend uses same dual-mode in
    `Config::read_secret("PDF_SERVICE_TOKEN")` so values stay in sync.

    #441 fail-closed: when a source is configured but resolves to empty (secret
    file unreadable / empty), we MUST NOT silently disable auth — ``INTERNAL_TOKEN``
    stays empty but ``source_configured`` is True, and `verify_internal_token`
    then rejects every request. Only a genuinely unconfigured service (dev/test,
    no source at all) passes through.
    """
    # env 變數「有設定但為空」也算 source configured（運維意圖啟用卻填空 → fail-closed），
    # 不可與「完全沒設」（None → dev/test）混為一談。
    raw_env_token = os.environ.get("PDF_SERVICE_TOKEN")
    env_source_configured = raw_env_token is not None
    token = raw_env_token.strip() if raw_env_token is not None else ""
    if token:
        return token, True
    file_path = os.environ.get("PDF_SERVICE_TOKEN_FILE", "").strip()
    if file_path:
        # 設定了 secret 檔路徑 = 預期啟用驗證；讀失敗 / 空檔 / 非 UTF-8 皆不可 fail-open。
        # UnicodeDecodeError 繼承 ValueError（非 OSError），須一併捕捉避免啟動 crash。
        try:
            content = Path(file_path).read_text(encoding="utf-8").strip()
        except (OSError, ValueError) as exc:
            log.error("PDF_SERVICE_TOKEN_FILE %s unreadable or invalid: %s", file_path, exc)
            content = ""
        return content, True
    return "", env_source_configured


INTERNAL_TOKEN, TOKEN_SOURCE_CONFIGURED = _load_internal_token()
if INTERNAL_TOKEN:
    log.info("X-Internal-Token validation ENABLED (token length=%d)", len(INTERNAL_TOKEN))
elif TOKEN_SOURCE_CONFIGURED:
    # #441：設定了 token 來源卻取不到值（prod secret 檔不可讀/空）→ 不可 fail-open。
    log.error(
        "X-Internal-Token source configured but resolved EMPTY — failing CLOSED "
        "(all render endpoints return 503 until the secret is fixed)"
    )
else:
    log.warning("X-Internal-Token validation DISABLED (no token source configured; dev/test)")


def verify_internal_token(
    x_internal_token: str | None = Header(default=None, alias="X-Internal-Token"),
) -> None:
    """FastAPI dependency: enforce X-Internal-Token on render endpoints.

    Three modes:
    - token set → enforce with `hmac.compare_digest` (constant-time).
    - source configured but token empty (#441 prod misconfig) → fail CLOSED (503).
    - no source configured (dev/test) → no-op pass-through.
    Applied to /render-* + /api/render only; /health* / /api/sample / /api/preview
    stay open (dev UI + container healthcheck).
    """
    if not INTERNAL_TOKEN:
        if TOKEN_SOURCE_CONFIGURED:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="PDF service token misconfigured (secret unreadable/empty)",
            )
        return
    if not x_internal_token or not hmac.compare_digest(x_internal_token, INTERNAL_TOKEN):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="invalid or missing X-Internal-Token",
        )


import tempfile

# Chromium (Playwright) 渲染引擎。取代 WeasyPrint：商用標楷體 DFKai-SB 為 point-matching
# composite，WeasyPrint 嵌入會破字；Chromium 原生處理（像瀏覽器/Word）→ 標楷體乾淨。
# 常駐 browser、每請求開新 page，效能與 WeasyPrint 同級。
_playwright = None
_browser = None


async def _start_browser() -> None:
    global _playwright, _browser
    from playwright.async_api import async_playwright

    _playwright = await async_playwright().start()
    _browser = await _playwright.chromium.launch(
        args=["--no-sandbox", "--allow-file-access-from-files", "--font-render-hinting=none"]
    )
    log.info("chromium browser launched")


async def _stop_browser() -> None:
    global _playwright, _browser
    try:
        if _browser is not None:
            await _browser.close()
        if _playwright is not None:
            await _playwright.stop()
    except Exception:  # noqa: BLE001
        log.exception("browser shutdown error (non-fatal)")


env = Environment(
    loader=FileSystemLoader(str(TEMPLATES_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
    trim_blocks=True,
    lstrip_blocks=True,
)


def _date_filter(v: Any, fmt: str = "%Y/%m/%d") -> str:
    if not v:
        return ""
    try:
        return v.strftime(fmt)
    except AttributeError:
        return str(v)


env.filters["date"] = _date_filter


class TemplateMeta(BaseModel):
    name: str
    schema_cls: type[BaseModel]
    html: str
    glp_doc_no: str = ""

    model_config = {"arbitrary_types_allowed": True}


TEMPLATES: dict[str, TemplateMeta] = {
    "pig_approval": TemplateMeta(name="IACUC 審查同意書", schema_cls=PigApprovalPayload, html="pig_approval.html"),
    "review_result": TemplateMeta(name="審核結果 (AD-04-01-10B)", schema_cls=ReviewResultPayload, html="review_result.html", glp_doc_no="AD-04-01-10B"),
    "review_reply": TemplateMeta(name="審查意見回覆表", schema_cls=ReviewReplyPayload, html="review_reply.html"),
    "aup_protocol": TemplateMeta(name="動物試驗研究計畫書 (R32)", schema_cls=AupProtocolPayload, html="aup_protocol.html", glp_doc_no="AD-04-01-01F"),
    "vet_patrol": TemplateMeta(name="欄位狀態表", schema_cls=VetPatrolPayload, html="vet_patrol.html", glp_doc_no="AD-05-01-02C"),
    "vet_patrol_report": TemplateMeta(name="獸醫巡場報告", schema_cls=VetPatrolReportPayload, html="vet_patrol_report.html"),
    "medical_record": TemplateMeta(name="實驗豬隻病歷總表", schema_cls=MedicalRecordPayload, html="medical_record.html"),
    "surgery": TemplateMeta(name="實驗豬隻手術紀錄表", schema_cls=SurgeryPayload, html="surgery.html"),
    "blood_test": TemplateMeta(name="動物血液檢查紀錄", schema_cls=BloodTestPayload, html="blood_test.html"),
    "warehouse": TemplateMeta(name="倉庫現況報表", schema_cls=WarehouseReportPayload, html="warehouse.html"),
    "audit_log": TemplateMeta(name="操作日誌", schema_cls=AuditLogPayload, html="audit_log.html"),
}


def _auto_fill_vet_patrol(data: dict[str, Any]) -> dict[str, Any]:
    """Fill server-side defaults: patrol_date (today) + period (AM/PM by hour).

    Defensive: 若 header 非 dict（或缺）→ 視為空 dict 後填預設。非 dict-truthy
    （如 list/str）→ ValueError 由 caller 轉 400（避免 schema validation 前 500）。
    """
    now = datetime.now(TZ)
    header = data.get("header")
    if header is None:
        header = {}
        data["header"] = header
    elif not isinstance(header, dict):
        raise ValueError("vet_patrol.header must be an object")
    if not header.get("patrol_date"):
        header["patrol_date"] = now.strftime("%Y-%m-%d")
    if not header.get("period"):
        header["period"] = "AM" if now.hour < 12 else "PM"
    return data


def _auto_fillers(template_id: str, data: dict[str, Any]) -> dict[str, Any]:
    if template_id == "vet_patrol":
        return _auto_fill_vet_patrol(data)
    return data


async def _warmup_render() -> None:
    """啟動暖機：跑一次極小的 HTML→PDF render，預先觸發 WeasyPrint 的 lazy import
    與 Pango/Cairo/fontconfig 字型快取建立（含 CJK glyph shaping），讓第一個真實
    請求不必承擔冷啟動成本。失敗為 non-fatal（僅記 log，不擋服務啟動）。
    """
    t0 = time.perf_counter()
    try:
        # 含中文字 → 一併暖 CJK 字型；走與正式渲染同一條 _render_pdf_async（過 render 信號量）。
        await _render_pdf_async("<html><body><p>暖機 warmup 預熱</p></body></html>")
        log.info("warmup render done in %.0f ms", (time.perf_counter() - t0) * 1000.0)
    except Exception:  # noqa: BLE001
        log.exception("warmup render failed (non-fatal)")


@asynccontextmanager
async def lifespan(_app: FastAPI):
    # 啟動常駐 Chromium（阻塞到 browser ready 才接請求，避免首個請求 race）。
    await _start_browser()
    # 背景觸發暖機（不阻塞服務接受連線）；以 PDF_WARMUP=0 關閉（如測試環境）。
    if os.environ.get("PDF_WARMUP", "1") != "0":
        asyncio.create_task(_warmup_render())
    else:
        log.info("warmup disabled (PDF_WARMUP=0)")
    yield
    await _stop_browser()


app = FastAPI(title="print-pdf", version="0.2.0", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/", response_class=HTMLResponse)
def index(_req: Request) -> HTMLResponse:
    return HTMLResponse((STATIC_DIR / "index.html").read_text(encoding="utf-8"))


@app.get("/healthz")
def healthz() -> dict[str, Any]:
    return {"status": "ok", "templates": sorted(TEMPLATES.keys())}


@app.get("/health")
def health() -> dict[str, Any]:
    """legacy compat — mirror of /healthz so backend probes pass during migration."""
    return healthz()


@app.get("/pdf-service-health")
def pdf_service_health() -> Response:
    """GLP 匯出前置存活探測。backend 代理此端點；服務有回應即代表可渲染 PDF。"""
    payload = {
        "service": "print-pdf",
        "engine": "chromium",
        "glp_ready": True,
    }
    return JSONResponse(payload, status_code=200)


@app.get("/api/templates")
def list_templates() -> dict[str, Any]:
    return {
        "templates": [
            {"id": k, "name": v.name, "glp_doc_no": v.glp_doc_no, "schema": v.schema_cls.model_json_schema()}
            for k, v in TEMPLATES.items()
        ]
    }


# ---------------------------------------------------------------------------
# Rendering primitives
# ---------------------------------------------------------------------------
def _render_html(template_id: str, data: dict[str, Any], extra_ctx: dict[str, Any] | None = None) -> str:
    if template_id not in TEMPLATES:
        raise HTTPException(404, f"Unknown template: {template_id}")
    meta = TEMPLATES[template_id]
    # auto_fillers + model_validate 都包進 try → 任何 shape 問題都走 400 而非 500
    try:
        data = _auto_fillers(template_id, data)
        payload = meta.schema_cls.model_validate(data)
    except (ValueError, TypeError) as exc:
        raise HTTPException(400, f"Invalid payload: {exc}") from exc
    except Exception as exc:
        raise HTTPException(400, f"Schema validation failed: {exc}") from exc
    tmpl = env.get_template(meta.html)
    ctx = payload.model_dump(mode="json")
    if extra_ctx:
        ctx.update(extra_ctx)
    return tmpl.render(**ctx)


def _aup_toc_pages(pdf_bytes: bytes) -> dict[str, int]:
    """兩遍渲染：Chromium 無 target-counter()，從第一遍 PDF 文字定位各章節起始頁。
    回傳 {sec1: 頁碼, ...}（1-based）。文字抽取失敗則回 {}（TOC 頁碼留白，不致命）。"""
    from io import BytesIO

    from pypdf import PdfReader

    titles = [
        ("sec1", "研究資料"), ("sec2", "計畫摘要"), ("sec3", "試驗物質與對照物質"),
        ("sec4", "研究設計與方法"), ("sec5", "相關規範及參考文獻"), ("sec6", "動物手術規劃"),
        ("sec7", "實驗動物資料"), ("sec8", "試驗人員資料"), ("sec9", "附件"),
    ]
    try:
        reader = PdfReader(BytesIO(pdf_bytes))
        texts = []
        for pg in reader.pages:
            try:
                texts.append((pg.extract_text() or "").replace(" ", "").replace("\n", ""))
            except Exception:  # noqa: BLE001
                texts.append("")
    except Exception:  # noqa: BLE001
        log.exception("aup toc: pdf parse failed")
        return {}
    result: dict[str, int] = {}
    start = 2  # 0-based：跳過封面(0)+目錄(1)
    for sec, title in titles:
        key = title.replace(" ", "")
        for i in range(start, len(texts)):
            if key in texts[i]:
                result[sec] = i + 1  # 1-based 頁碼
                start = i  # 章節依序遞增，下個從此頁起找
                break
    return result


async def _two_pass_aup_pdf(render_html) -> bytes:
    """計畫書專用：第一遍空頁碼→定位章節頁→第二遍回填 TOC。render_html(toc_pages)->html。"""
    pdf1 = await _render_pdf_async(render_html({}))
    toc = _aup_toc_pages(pdf1)
    if not toc:
        return pdf1
    return await _render_pdf_async(render_html(toc))


def _render_html_from_payload(template_id: str, payload: BaseModel, extra_ctx: dict[str, Any] | None = None) -> str:
    meta = TEMPLATES[template_id]
    tmpl = env.get_template(meta.html)
    ctx = payload.model_dump(mode="json")
    if extra_ctx:
        ctx.update(extra_ctx)
    return tmpl.render(**ctx)


async def _html_to_pdf_async(html: str) -> bytes:
    """HTML→PDF via Chromium。temp file 寫在 BASE(/app) → 模板 @font-face 的
    file:///app/fonts/* 絕對路徑可解析（goto file:// 同源 + --allow-file-access-from-files）。
    prefer_css_page_size：採用 @page size（A4 / A4 landscape）；print_background：保留 th 灰底等。"""
    fd, path = tempfile.mkstemp(suffix=".html", dir=str(BASE))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(html)
        page = await _browser.new_page()
        try:
            t0 = time.perf_counter()
            await page.goto("file://" + path, wait_until="networkidle")
            # 等字型載入完 + 任何 JS 自適應（如 vet_patrol 儲存格縮放）跑完才印。
            # 無 pen-grid 的文件條件立即為真、不受影響；逾時不致命。
            try:
                await page.wait_for_function(
                    "(!document.fonts || document.fonts.status === 'loaded') && "
                    "(!document.querySelector('table.pen-grid') || "
                    "document.documentElement.getAttribute('data-fit-done') === '1')",
                    timeout=5000,
                )
            except Exception:  # noqa: BLE001
                log.warning("wait_for_function (fonts/fit) timed out; proceeding")
            pdf = await page.pdf(prefer_css_page_size=True, print_background=True)
            logging.info("chromium render: %.0f ms, %d bytes", (time.perf_counter() - t0) * 1000.0, len(pdf))
            return pdf
        finally:
            await page.close()
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


# vet_patrol_report 專用 header/footer template。
#
# 不用 @font-face + file:///app/fonts/kaiu.ttf（主頁排版那套做法）：Chromium 把
# headerTemplate/footerTemplate 渲染在獨立、受限的 iframe，@font-face 的 file://
# 資源常常載入不到（Puppeteer/Playwright 已知限制）。因此頁首頁尾字型必須是**系統
# 註冊字型**（fontconfig）才引用得到。
#
# 使用者要求頁首頁尾與內文一致（標楷體 + Times New Roman，非 Noto）。故 Dockerfile
# 已把 kaiu.ttf(標楷體) + times*.ttf(Times New Roman) 裝進 /usr/share/fonts + fc-cache，
# 這裡即可用 family name 引用。字型序：Times New Roman 先（管英數 / 文件編號代碼 /
# 頁碼 / of），CJK 逐字 fallback 到「標楷體」。
# 註：CJK 名用「標楷體」而非「DFKai-SB」——後者被容器 fontconfig 強別名綁到 WenQuanYi，
# 「標楷體」才解析到真 kaiu.ttf（實測 fc-match 確認）。
#
# width:100% 外還加 margin 左右留白會讓 flex 容器超出 Chromium 保留的 header/footer
# 版面，把靠右對齊的頁碼擠出畫面；改用 box-sizing:border-box + padding，留白內含
# 在 100% 寬度裡，不會溢出。
_VET_PATROL_REPORT_HEADER_HTML = (
    '<div style="box-sizing:border-box; width:100%; padding:0 10mm; font-size:9pt; '
    'display:flex; justify-content:space-between; '
    'font-family:\'Times New Roman\',\'標楷體\',serif; color:#000;">'
    '<span>文件編號 AD-02-02-01</span>'
    '<span>頁次/總頁數 <span class="pageNumber"></span> of <span class="totalPages"></span></span>'
    '</div>'
)

_VET_PATROL_REPORT_FOOTER_HTML = (
    '<div style="box-sizing:border-box; width:100%; padding:0 10mm; text-align:center; '
    'font-size:9pt; line-height:1.4; '
    'font-family:\'Times New Roman\',\'標楷體\',serif; color:#000;">'
    '<div>版權為豬博士動物科技股份有限公司所有，禁止任何未經授權的使用</div>'
    '<div>All Rights Reserved &#169; DrPIG. Unauthorized use in any form is prohibited.</div>'
    '</div>'
)

_VET_PATROL_REPORT_MARGIN = {"top": "16mm", "bottom": "16mm", "left": "10mm", "right": "10mm"}


async def _html_to_pdf_paginated_async(
    html: str, header_html: str, footer_html: str, margin: dict[str, str]
) -> bytes:
    """HTML→PDF via Chromium，帶 Playwright 原生 header/footer template（每頁自動重複 + 自動頁碼）。

    base.html 的 CSS Paged Media（@top-left / @bottom-center）是 WeasyPrint 專屬語法，
    Chromium 的 print-to-pdf 引擎不支援、會被靜默忽略。vet_patrol_report 從固定 mm
    絕對定位（每個 .sheet 一頁）改回一般文件流排版後，需要跨頁重複頁首/頁尾，
    改走 Playwright `page.pdf(display_header_footer=...)` 這條原生機制。
    """
    fd, path = tempfile.mkstemp(suffix=".html", dir=str(BASE))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(html)
        page = await _browser.new_page()
        try:
            t0 = time.perf_counter()
            await page.goto("file://" + path, wait_until="networkidle")
            # 等字型載入完 + 任何 JS 自適應跑完才印（與 _html_to_pdf_async 同樣條件）。
            # vet_patrol_report 的一頁化壓縮必須在 page.pdf() 之前完成，否則印到的是
            # 未壓縮的版面；只等 fonts 會有 race（fonts.ready 的 callback 尚未執行完）。
            # 無 .autofit-root 的文件條件立即為真、不受影響；逾時不致命。
            try:
                await page.wait_for_function(
                    "(!document.fonts || document.fonts.status === 'loaded') && "
                    "(!document.querySelector('.autofit-root') || "
                    "document.documentElement.getAttribute('data-fit-done') === '1')",
                    timeout=5000,
                )
            except Exception:  # noqa: BLE001
                log.warning("wait_for_function (fonts/fit) timed out; proceeding")
            pdf = await page.pdf(
                prefer_css_page_size=True,
                print_background=True,
                display_header_footer=True,
                header_template=header_html,
                footer_template=footer_html,
                margin=margin,
            )
            logging.info(
                "chromium paginated render: %.0f ms, %d bytes",
                (time.perf_counter() - t0) * 1000.0, len(pdf),
            )
            return pdf
        finally:
            await page.close()
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


# 並行上限：Chromium 每 page 為獨立 renderer process（thread-safe，不像 WeasyPrint
# 的 Pango/Cairo 會 core dump）。預設 2 兼顧吞吐與筆電記憶體；可經 PDF_RENDER_CONCURRENCY 覆寫。
def _render_concurrency() -> int:
    try:
        n = int(os.environ.get("PDF_RENDER_CONCURRENCY", "2"))
    except ValueError:
        n = 2
    return max(1, n)


_RENDER_SEMAPHORE = asyncio.Semaphore(_render_concurrency())


async def _render_pdf_async(html: str) -> bytes:
    """HTML→PDF（Chromium），由全域信號量節流以控記憶體。"""
    async with _RENDER_SEMAPHORE:
        return await _html_to_pdf_async(html)


async def _render_pdf_paginated_async(
    html: str, header_html: str, footer_html: str, margin: dict[str, str]
) -> bytes:
    """HTML→PDF（Chromium，帶原生 header/footer template），同一信號量節流。"""
    async with _RENDER_SEMAPHORE:
        return await _html_to_pdf_paginated_async(html, header_html, footer_html, margin)


def _pdf_response(pdf: bytes, filename: str) -> Response:
    from urllib.parse import quote
    ascii_name = filename.encode("ascii", "replace").decode()
    utf8_name = quote(filename)
    return Response(
        pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"inline; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}",
            "X-PDF-Renderer": "chromium",
        },
    )


# ---------------------------------------------------------------------------
# Generic endpoints (/api/render/{id})
# ---------------------------------------------------------------------------
@app.post("/api/render/{template_id}", dependencies=[Depends(verify_internal_token)])
async def render(
    template_id: str,
    payload: dict[str, Any],
    output_format: str = Query("pdf", alias="format"),
) -> Response:
    html = _render_html(template_id, payload)
    if output_format == "html":
        return HTMLResponse(html)
    try:
        if template_id == "aup_protocol":
            pdf = await _two_pass_aup_pdf(
                lambda tp: _render_html(template_id, payload, extra_ctx={"toc_pages": tp})
            )
        else:
            pdf = await _render_pdf_async(html)
    except Exception as exc:  # noqa: BLE001
        log.exception("PDF render failed for %s", template_id)
        raise HTTPException(500, f"PDF engine error: {exc}") from exc
    return _pdf_response(pdf, f"{template_id}.pdf")


@app.get("/api/sample/{template_id}")
def sample(template_id: str) -> dict[str, Any]:
    if template_id not in SAMPLES:
        raise HTTPException(404, f"No sample data for template: {template_id}")
    meta = TEMPLATES[template_id]
    return meta.schema_cls.model_validate(SAMPLES[template_id]).model_dump(mode="json")


@app.get("/api/preview/{template_id}")
async def preview(
    template_id: str,
    output_format: str = Query("html", alias="format"),
) -> Response:
    if template_id not in SAMPLES:
        raise HTTPException(404, f"No sample for template: {template_id}")
    html = _render_html(template_id, SAMPLES[template_id])
    if output_format == "html":
        return HTMLResponse(html)
    if template_id == "aup_protocol":
        pdf = await _two_pass_aup_pdf(
            lambda tp: _render_html(template_id, SAMPLES[template_id], extra_ctx={"toc_pages": tp})
        )
    else:
        pdf = await _render_pdf_async(html)
    return _pdf_response(pdf, f"{template_id}.pdf")


# ---------------------------------------------------------------------------
# Adapter-aware endpoints (drop-in replacement for pdf-service routes)
# ---------------------------------------------------------------------------
def _adapter_call(adapter_fn, body: dict[str, Any], label: str) -> BaseModel:
    try:
        return adapter_fn(body)
    except (ValueError, TypeError) as e:
        raise HTTPException(400, f"{label} adapter: {e}") from e
    except Exception as e:  # noqa: BLE001
        log.exception("%s adapter failed", label)
        raise HTTPException(400, f"{label} adapter failed: {type(e).__name__}") from e


def _require_dict(body: Any) -> dict[str, Any]:
    if not isinstance(body, dict):
        raise HTTPException(400, "Body must be a JSON object")
    return body


@app.post("/render-aup/from-working-content", dependencies=[Depends(verify_internal_token)])
async def render_aup_from_working_content(
    request: Request,
    output_format: str = Query("pdf", alias="format"),
) -> Response:
    body = _require_dict(await request.json())
    wc = body.get("working_content")
    if not isinstance(wc, dict):
        raise HTTPException(400, 'Body must be {"working_content": {...}}')
    payload = _adapter_call(aup_ad.from_working_content, wc, "aup_protocol")
    # format=html：回傳送進渲染前的同一份 HTML，供前端預覽 iframe，
    # 確保「計畫內容」預覽與匯出 PDF 同源（同模板 + 同資料）一致。
    if output_format == "html":
        return HTMLResponse(_render_html_from_payload("aup_protocol", payload))
    # 兩遍渲染回填目錄頁碼（Chromium 無 target-counter）
    pdf = await _two_pass_aup_pdf(
        lambda tp: _render_html_from_payload("aup_protocol", payload, extra_ctx={"toc_pages": tp})
    )
    no = payload.protocol.iacuc_apply_no or payload.protocol.iacuc_approval_no or "draft"
    return _pdf_response(pdf, f"AUP_{no}.pdf")


@app.post("/render-review-reply/from-review-data", dependencies=[Depends(verify_internal_token)])
async def render_review_reply_from_review_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(review_reply_ad.from_review_reply_data, body, "review_reply")
    html = _render_html_from_payload("review_reply", payload)
    pdf = await _render_pdf_async(html)
    return _pdf_response(pdf, f"review_reply_{payload.application_no or 'draft'}.pdf")


@app.post("/render-review-result/from-review-data", dependencies=[Depends(verify_internal_token)])
async def render_review_result_from_review_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(review_result_ad.from_review_data, body, "review_result")
    html = _render_html_from_payload("review_result", payload)
    pdf = await _render_pdf_async(html)
    iacuc = (body.get("protocol") or {}).get("iacuc_no") or "draft"
    return _pdf_response(pdf, f"review_result_{iacuc}.pdf")


@app.post("/render-medical-record/from-animal-data", dependencies=[Depends(verify_internal_token)])
async def render_medical_record_from_animal_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(medical_record_ad.from_animal_data, body, "medical_record")
    html = _render_html_from_payload("medical_record", payload)
    pdf = await _render_pdf_async(html)
    iacuc = payload.animal.iacuc_no or "unassigned"
    ear = payload.animal.ear_tag or "draft"
    return _pdf_response(pdf, f"medical_record_{iacuc}_{ear}.pdf")


@app.post("/render-surgery/from-surgery-data", dependencies=[Depends(verify_internal_token)])
async def render_surgery_from_surgery_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(surgery_ad.from_surgery_data, body, "surgery")
    html = _render_html_from_payload("surgery", payload)
    pdf = await _render_pdf_async(html)
    iacuc = payload.animal.iacuc_no or "unassigned"
    ear = payload.animal.ear_tag or "draft"
    date = payload.surgery.surgery_date or "draft"
    return _pdf_response(pdf, f"surgery_{iacuc}_{ear}_{date}.pdf")


@app.post("/render-blood-test/from-blood-test-data", dependencies=[Depends(verify_internal_token)])
async def render_blood_test_from_blood_test_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(blood_test_ad.from_blood_test_data, body, "blood_test")
    html = _render_html_from_payload("blood_test", payload)
    pdf = await _render_pdf_async(html)
    iacuc = payload.animal_iacuc_no or "unassigned"
    ear = payload.animal_ear_tag or "draft"
    return _pdf_response(pdf, f"blood_test_{iacuc}_{ear}.pdf")


@app.post("/render-audit-log/from-export-data", dependencies=[Depends(verify_internal_token)])
async def render_audit_log_from_export_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(audit_log_ad.from_export_data, body, "audit_log")
    html = _render_html_from_payload("audit_log", payload)
    pdf = await _render_pdf_async(html)
    p_from = (payload.meta.period_from or "all").replace(":", "").replace(" ", "_")
    p_to = (payload.meta.period_to or "all").replace(":", "").replace(" ", "_")
    return _pdf_response(pdf, f"audit_log_{p_from}_{p_to}.pdf")


@app.post("/render-warehouse/from-report-data", dependencies=[Depends(verify_internal_token)])
async def render_warehouse_from_report_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(warehouse_ad.from_warehouse_report, body, "warehouse")
    # 倉庫平面圖 PNG → data URI 嵌入 HTML
    try:
        png_bytes = render_floor_plan_png(body.get("locations") or [])
    except Exception:  # noqa: BLE001
        log.exception("floor_plan PNG render failed")
        png_bytes = b""
    extra = {
        "floor_plan_data_uri": (
            "data:image/png;base64," + base64.b64encode(png_bytes).decode("ascii")
            if png_bytes else ""
        ),
    }
    html = _render_html_from_payload("warehouse", payload, extra_ctx=extra)
    pdf = await _render_pdf_async(html)
    return _pdf_response(pdf, f"warehouse_{payload.warehouse.code or 'draft'}.pdf")


@app.post("/render-vet-patrol/from-animals", dependencies=[Depends(verify_internal_token)])
async def render_vet_patrol_from_animals(request: Request) -> Response:
    body = _require_dict(await request.json())
    animals = body.get("animals")
    if not isinstance(animals, list):
        raise HTTPException(400, "`animals` must be a list")
    try:
        payload = vet_patrol_ad.from_animals(
            animals,
            inspector_name=body.get("inspector_name", ""),
            patrol_date=body.get("patrol_date", ""),
            period=body.get("period", ""),
        )
    except Exception as e:  # noqa: BLE001
        log.exception("vet_patrol adapter failed")
        raise HTTPException(400, "Adapter failed") from e
    # 自動補日期 / 時段
    data = payload.model_dump(mode="json")
    _auto_fill_vet_patrol(data)
    html = _render_html("vet_patrol", data)
    pdf = await _render_pdf_async(html)
    return _pdf_response(pdf, f"vet_patrol_{data['header'].get('patrol_date','draft')}.pdf")


@app.post("/render-vet-patrol-report/from-report-data", dependencies=[Depends(verify_internal_token)])
async def render_vet_patrol_report_from_report_data(request: Request) -> Response:
    body = _require_dict(await request.json())
    payload = _adapter_call(vet_patrol_report_ad.from_report_data, body, "vet_patrol_report")
    html = _render_html_from_payload("vet_patrol_report", payload)
    pdf = await _render_pdf_paginated_async(
        html, _VET_PATROL_REPORT_HEADER_HTML, _VET_PATROL_REPORT_FOOTER_HTML, _VET_PATROL_REPORT_MARGIN
    )
    date = (payload.patrol_date or "draft").replace("-", "")
    return _pdf_response(pdf, f"試驗豬場巡場報告_{date}.pdf")


# ---------------------------------------------------------------------------
# project_medical：N 隻動物 → 1 個合併 PDF（並行渲染 + pypdf merge）
# ---------------------------------------------------------------------------
@app.post("/render-project-medical/from-project-data", dependencies=[Depends(verify_internal_token)])
async def render_project_medical_from_project_data(request: Request) -> Response:
    """N 隻動物各自渲染 medical_record HTML → PDF，再用 pypdf 合併保序回傳。"""
    from io import BytesIO

    from pypdf import PdfReader, PdfWriter

    body = _require_dict(await request.json())
    animals = body.get("animals")
    if not isinstance(animals, list) or not animals:
        raise HTTPException(400, "`animals` must be a non-empty list")
    iacuc_no = body.get("iacuc_no") or "draft"

    sem = asyncio.Semaphore(4)  # 限制並行避免記憶體爆

    async def _one(idx: int, animal_data: Any) -> bytes | None:
        if not isinstance(animal_data, dict):
            return None
        try:
            payload = medical_record_ad.from_animal_data(animal_data)
        except (ValueError, TypeError) as e:
            log.warning("project_medical[%d] adapter failed: %s", idx, e)
            return None
        async with sem:
            try:
                html = _render_html_from_payload("medical_record", payload)
                return await _render_pdf_async(html)
            except Exception:
                log.exception("project_medical[%d] render failed", idx)
                return None

    results = await asyncio.gather(*(_one(i, a) for i, a in enumerate(animals)))

    writer = PdfWriter()
    failed: list[int] = []
    for idx, pdf in enumerate(results):
        if pdf is None:
            failed.append(idx)
            continue
        try:
            for page in PdfReader(BytesIO(pdf)).pages:
                writer.add_page(page)
        except Exception:
            log.exception("project_medical[%d] pypdf merge failed", idx)
            failed.append(idx)

    if len(writer.pages) == 0:
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            f"No pages produced — all {len(animals)} animals failed",
        )

    if failed:
        log.warning("project_medical: %d/%d animals failed: %s", len(failed), len(animals), failed)

    out = BytesIO()
    writer.write(out)
    return _pdf_response(out.getvalue(), f"project_medical_{iacuc_no}.pdf")


# ─── R53-10: Weekly medical report xlsx ─────────────────────────────
@app.post("/render-xlsx/weekly-medical-report", dependencies=[Depends(verify_internal_token)])
async def render_weekly_medical_report_xlsx(request: Request) -> Response:
    """接收 MedicalTimelineEvent[] JSON，產出對齊範本的 10 欄 xlsx。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    body = _require_dict(await request.json())
    events = body.get("events")
    if not isinstance(events, list):
        raise HTTPException(400, "`events` must be a list")

    def _safe(v: Any) -> str:
        s = str(v) if v is not None else ""
        if s and s[0] in ("=", "+", "-", "@"):
            return "'" + s
        return s

    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"

    headers = ["日期", "耳號", "出生日期", "體重", "試驗單位", "試驗內容",
               "特殊儀器使用", "麻醉總時數", "麻醉開始時間", "麻醉結束時間"]
    col_widths = [14, 8, 14, 8, 30, 60, 20, 10, 12, 12]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, ev in enumerate(events, 2):
        event_date = ev.get("event_date", "")
        birth_date = ev.get("birth_date", "")
        ws.cell(row=row_idx, column=1, value=event_date).border = thin_border
        ws.cell(row=row_idx, column=2, value=_safe(ev.get("ear_tag", ""))).border = thin_border
        ws.cell(row=row_idx, column=3, value=birth_date).border = thin_border
        ws.cell(row=row_idx, column=4, value=ev.get("latest_weight")).border = thin_border

        protocol = ev.get("protocol_title") or ev.get("iacuc_no") or ""
        ws.cell(row=row_idx, column=5, value=_safe(protocol)).border = thin_border

        details = ev.get("details") or ev.get("summary", "")
        ws.cell(row=row_idx, column=6, value=_safe(details)).border = thin_border
        ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True)

        ws.cell(row=row_idx, column=7, value=_safe(ev.get("equipment_used") or "")).border = thin_border

        start = ev.get("anesthesia_start")
        end = ev.get("anesthesia_end")
        if start and end:
            try:
                from datetime import datetime as dt
                t_start = dt.fromisoformat(str(start).replace("Z", "+00:00")).astimezone(TZ)
                t_end = dt.fromisoformat(str(end).replace("Z", "+00:00")).astimezone(TZ)
                delta = t_end - t_start
                total_min = int(delta.total_seconds() // 60)
                h, m = divmod(total_min, 60)
                duration_str = f"{h}hr {m:02d}min" if h else f"{m}min"
                ws.cell(row=row_idx, column=8, value=duration_str).border = thin_border
                ws.cell(row=row_idx, column=9, value=t_start.strftime("%H:%M")).border = thin_border
                ws.cell(row=row_idx, column=10, value=t_end.strftime("%H:%M")).border = thin_border
            except (ValueError, TypeError):
                for c in (8, 9, 10):
                    ws.cell(row=row_idx, column=c, value="").border = thin_border
        else:
            for c in (8, 9, 10):
                ws.cell(row=row_idx, column=c, value="").border = thin_border

    ws.auto_filter.ref = f"A1:J{len(events) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="weekly_medical_report.xlsx"'},
    )


# ─── R53-11: Weekly medical report PDF ────────────────────────────────
@app.post("/render-pdf/weekly-medical-report", dependencies=[Depends(verify_internal_token)])
async def render_weekly_medical_report_pdf(request: Request) -> Response:
    """接收 MedicalTimelineEvent[] JSON，產出橫式 A4 PDF。"""
    from datetime import datetime as dt

    body = _require_dict(await request.json())
    events = body.get("events")
    if not isinstance(events, list):
        raise HTTPException(400, "`events` must be a list")

    for ev in events:
        start = ev.get("anesthesia_start")
        end = ev.get("anesthesia_end")
        if start and end:
            try:
                t_start = dt.fromisoformat(str(start).replace("Z", "+00:00")).astimezone(TZ)
                t_end = dt.fromisoformat(str(end).replace("Z", "+00:00")).astimezone(TZ)
                delta = t_end - t_start
                total_min = int(delta.total_seconds() // 60)
                h, m = divmod(total_min, 60)
                ev["anesthesia_duration"] = f"{h}hr {m:02d}min" if h else f"{m}min"
                ev["anesthesia_start_display"] = t_start.strftime("%H:%M")
                ev["anesthesia_end_display"] = t_end.strftime("%H:%M")
            except (ValueError, TypeError):
                pass

    date_range = body.get("date_range", "")
    html = _render_html("weekly_medical_report", {"events": events, "date_range": date_range})
    pdf = await _render_pdf_async(html)
    return _pdf_response(pdf, "weekly_medical_report.pdf")


# ─── R53-15: Byproduct monthly report xlsx ──────────────────────────
@app.post("/render-xlsx/byproduct-monthly", dependencies=[Depends(verify_internal_token)])
async def render_byproduct_monthly_xlsx(request: Request) -> Response:
    """接收 ByproductMonthlyRow[] JSON，產出 6 欄月結 xlsx。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    body = await request.json()
    rows = body.get("rows")
    if not isinstance(rows, list):
        raise HTTPException(400, "`rows` must be a list")

    wb = Workbook()
    ws = wb.active
    ws.title = "月結報表"

    headers = ["採樣日期", "案子", "耳號", "需求客戶", "採樣內容", "記錄者"]
    col_widths = [14, 30, 8, 25, 50, 12]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, r in enumerate(rows, 2):
        sampled = r.get("sampled_at", "")
        if sampled:
            try:
                from datetime import datetime as dt
                parsed = dt.fromisoformat(str(sampled).replace("Z", "+00:00"))
                sampled = parsed.astimezone(TZ).strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                if "T" in str(sampled):
                    sampled = str(sampled).split("T")[0]
        protocol = r.get("protocol_title") or r.get("iacuc_no") or ""
        ws.cell(row=row_idx, column=1, value=sampled).border = thin_border
        ws.cell(row=row_idx, column=2, value=protocol).border = thin_border
        ws.cell(row=row_idx, column=3, value=r.get("ear_tag", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=r.get("requester_display") or "").border = thin_border
        ws.cell(row=row_idx, column=5, value=r.get("sample_content", "")).border = thin_border
        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=6, value=r.get("collector_name") or "").border = thin_border

    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="byproduct_monthly_report.xlsx"'},
    )
