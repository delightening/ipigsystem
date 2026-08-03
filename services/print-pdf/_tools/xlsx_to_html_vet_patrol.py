"""One-shot structural analysis: dump xlsx vet_patrol_template layout.

⚠️ 不是 codegen。final `templates/vet_patrol.html` 已手動重構為 `{% for %}` loops
提升可讀性，重跑此工具不會也不應直接覆蓋。本工具的用途：

  - vet/QA 改 xlsx 後，dev 跑此工具看 cell 位置 / merge 範圍 / col widths
    的 delta，再手動把改動同步到 `vet_patrol.html`。
  - debug grid 結構（merge 範圍對不上、colspan 算錯等）。

Why bother: reference PDF (動物欄位巡視報告範例.pdf) 是 vet/QA 手調 2D spatial
layout（A/B/C/D/E/F/G 區位置標籤散落、merge cells、空白列保留），xlsx 本身已
是 Jinja2 template（cells 含 `{{ pens.D18.ear_tags }}`）。手寫 HTML 容易漏 merge
/ 算錯 colspan，用工具掃過 xlsx 比對比較不會出錯。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# repo-relative：services/print-pdf/_tools/xxx.py → ../../../templates/vet_patrol_template.xlsx
REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
XLSX = REPO_ROOT / "templates" / "vet_patrol_template.xlsx"
if not XLSX.is_file():
    raise SystemExit(f"xlsx not found: {XLSX}")

wb = openpyxl.load_workbook(XLSX, data_only=False)
ws = wb.active
rows, cols = ws.max_row, ws.max_column

# build "occupied by merge" lookup; secondary cells skipped in output
merge_master: dict[tuple[int, int], tuple[int, int]] = {}  # (r,c) → (rowspan, colspan)
merge_skip: set[tuple[int, int]] = set()
for mr in ws.merged_cells.ranges:
    rs = mr.max_row - mr.min_row + 1
    cs = mr.max_col - mr.min_col + 1
    merge_master[(mr.min_row, mr.min_col)] = (rs, cs)
    for r in range(mr.min_row, mr.max_row + 1):
        for c in range(mr.min_col, mr.max_col + 1):
            if (r, c) != (mr.min_row, mr.min_col):
                merge_skip.add((r, c))


def classify(value) -> str:
    """Return CSS class for cell based on value pattern."""
    if value is None:
        return "empty"
    s = str(value)
    if "{{ pens." in s and ".status }}" in s:
        return "status-cell"
    if "{{ pens." in s and ".ear_tags" in s:
        return "tag-cell"
    if s in ("○", "●"):
        return "status-cell"
    if "{{ header." in s:
        return "header-input"
    # Zone labels (single uppercase letter A-G)
    if len(s) == 1 and s in "ABCDEFG":
        return "zone-label"
    if "巡視" in s:
        return "header-label"
    if s.startswith("□") or "全場" in s:
        return "footer"
    if s in ("AM", "PM"):
        return "ampm"
    return "text"


# column width: xlsx char unit → %. Total = sum of widths
col_widths = []
DEFAULT_WIDTH = 8.43
for c in range(1, cols + 1):
    letter = get_column_letter(c)
    w = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
    col_widths.append(w or DEFAULT_WIDTH)
total_w = sum(col_widths)
col_pct = [w / total_w * 100 for w in col_widths]

print(f"<!-- STRUCTURAL ANALYSIS dump from {XLSX.name} -->")
print("<!-- Reference only; do NOT copy verbatim into vet_patrol.html (which uses {% for %} loops). -->")
print("<!-- Use to spot delta when vet/QA edits xlsx; manually sync changes into the loop tuples. -->")
print()
print('<table class="pen-grid">')
print("  <colgroup>")
for pct in col_pct:
    print(f'    <col style="width: {pct:.2f}%">')
print("  </colgroup>")
print("  <tbody>")
for r in range(1, rows + 1):
    print("    <tr>")
    for c in range(1, cols + 1):
        if (r, c) in merge_skip:
            continue
        cell = ws.cell(row=r, column=c)
        rowspan, colspan = merge_master.get((r, c), (1, 1))
        klass = classify(cell.value)
        attrs = []
        if rowspan > 1:
            attrs.append(f'rowspan="{rowspan}"')
        if colspan > 1:
            attrs.append(f'colspan="{colspan}"')
        if klass:
            attrs.append(f'class="{klass}"')
        attr_str = (" " + " ".join(attrs)) if attrs else ""
        # render value; preserve {{ }} as Jinja
        val = "" if cell.value is None else str(cell.value)
        # for status-cell with literal ○/●, keep as-is. For Jinja vars, keep as-is too.
        print(f'      <td{attr_str}>{val}</td>')
    print("    </tr>")
print("  </tbody>")
print("</table>")
